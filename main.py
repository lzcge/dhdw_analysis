import time
import file_deal
import linecache
import os
from openpyxl import Workbook
import re


def sava_data_xlsx(ps_info_dict: dict):
    '''
    结果写入文件
    :param ps_info_list:文件对象list
    :return:
    '''
    # 新建工作簿
    wb = Workbook()
    for file_path_name,ps_file_info_list in ps_info_dict.items():
        print("正在将以下文件中的分析结果写入结果文件：")
        print(file_path_name)
        # 每个ps文件数据创建一个sheet工作表保存
        ps_name = '_'.join((file_path_name.split('\\')[-1]).split("_")[:2])
        ws = wb.create_sheet(ps_name,-1)
        # 最上面先写入一行这个ps文件的路径，标识这些数据的来源文件
        c = ws.cell(row=1, column=1)
        c.value = file_path_name

        ws.append(["burst_type","sub_chan","fn","tsn","beamid","bandid","crc","dl_caridx"])
        # 如果结果数据列表不为空就写入数据
        # 为空就写入无数据结果
        if ps_file_info_list:
            for info in ps_file_info_list:
            # for infoclass in (file.info_list).values():
                ws.append(info)
        else:
            ws.append(["未查找到相关数据"])

    # 获取当前时间戳
    timestamp = str(int(time.time()))
    # 结果文件保存路径
    result_dir = os.getcwd()+'\\'+'result'
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    result_file_path = os.path.join(os.getcwd()+r'\result','result_'+timestamp+'.xlsx')
    # 如果结果文件存在，将原来相同名称的文件通过时间戳进行改名备份，然后再新建文件
    # 不存在就直接新建
    if os.path.exists(result_file_path):
        os.rename(result_file_path, result_file_path + '_bak_' + timestamp)  # 将原文件重命名备份
    # 保存最终文件
    wb.save(result_file_path)
    print("*****文件分析完成*****")
    print("结果文件路径：" + result_file_path)


def find_ps_files_list(file_path: str) -> list:
    '''
    查找目标路径下包含的所有的ps文件
    :param file_path:
    :return:
    '''
    walk_generator = os.walk(file_path)
    file_root_path = {}
    ps_paths = []
    # 查看出目标路径下的所有文件和对应的上级路径，得到 {“文件夹路径”：“文件名”}
    for root_path, dirs, all_files in walk_generator:
        if len(all_files) > 0:
            file_root_path[root_path] = all_files

    # 根据得到的{“文件夹路径”：“文件名”}，查找出每个文件夹路径下的ps文件，并拼接成完整文件路径返回 [“ps_file_path”]
    for key,val in file_root_path.items():
        ps_files = [file for file in val if "_ps.dat" in file and "_ps.dat." not in file ]
        if ps_files:
            file_path_name = os.path.join(key,ps_files[0])
            ps_paths.append(file_path_name)
            # ps_paths.append(os.path.join(key,ps_files[0]))

    return ps_paths


def data_deal(ps_files: list):
    '''
    数据查找和逻辑处理
    :param ps_files:
    :param pattern_rule:
    :return:
    '''

    psfile_result_dict={}
    # 遍历每个ps文件去找到每个ps文件中的所有信息
    for ps_file_path in ps_files:
        ps_info_list = []  # 存储最终结果：File对象列表
        # 获取当前文件对象的文件名称路径
        # ps_file_path = file.file_path_name
        print("正在分析文件：")
        print(ps_file_path)
        # 获取该文件中所有符合所有匹配规则的行，做一次初筛，将不需要的以及异常的行数据踢除，减少对异常日志的异常处理，以及提升后续逻辑查找处理的速度
        contents = linecache.getlines(ps_file_path)
        ps_pattern_contents = [line for line in contents if "L1C_L1A_RX_DATA_IND (BURST_TYPE_PMBCH,SUB_CHAN_PMBCH_DH" in line or "set DL schedule[1]: BURST_TYPE_PMBCH,SUB_CHAN_PMBCH_DH" in line]

        # 根据关键词查找获取相关信息
        dl_fn,dl_bandid,dl_tsn,dl_caridx=-1,-1,-1,-1
        for i,line in enumerate(ps_pattern_contents):
            if "set DL schedule[1]: BURST_TYPE_PMBCH,SUB_CHAN_PMBCH_DH" in line:
                # 提取fn
                fn_match = re.findall(r'time\[(\d+)', line)
                dl_fn = int(fn_match[0]) if fn_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取tsn
                tsn_match = re.findall(r'time\[\d+,(\d+)', line)
                dl_tsn = int(tsn_match[0]) if tsn_match else -999999

                # 提取 carIdx
                caridx_match = re.findall(r'carIdx\((\d+)', line)
                dl_caridx = int(caridx_match[0]) if caridx_match else -999999

            # if "set DL schedule[1]: band" in line:
            #     bandid_match = re.findall(r'band\s+(\d+)', line)
            #     dl_bandid = int(bandid_match[0]) if bandid_match else -999999

            if "L1C_L1A_RX_DATA_IND" in line and "BURST_TYPE_PMBCH,SUB_CHAN_PMBCH_DH" in line:
                # 提取BURST_TYPE
                burst_type_match = re.findall(r'BURST_TYPE_(\w+)', line)
                burst_type = burst_type_match[0] if burst_type_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取SUB_CHAN
                sub_chan_match = re.findall(r'SUB_CHAN_(\w+)', line)
                sub_chan = sub_chan_match[0] if sub_chan_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取fn
                fn_match = re.findall(r'fn=(\d+)', line)
                fn = int(fn_match[0]) if fn_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取tsn
                tsn_match = re.findall(r'tsn=(\d+)', line)
                tsn = int(tsn_match[0]) if tsn_match else -999999

                # 提取 beamid
                beamid_match = re.findall(r'beamId=(\d+)', line)
                beamid = int(beamid_match[0]) if beamid_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取 bandId
                bandid_match = re.findall(r'bandId=(\d+)', line)
                bandid = int(bandid_match[0]) if bandid_match else -999999  # 如果未找到，赋值-999999的异常值

                # 提取 crc
                crc_match = re.findall(r'crc=(\d+)', line)
                crc = int(crc_match[0]) if crc_match else -999999

                # # 提取 rssi
                # rssi_match = re.findall(r'rssi=(-?\d+)', line)
                # rssi = int(rssi_match[0]) if rssi_match else -999999  # 如果未找到，赋值-999999的异常值
                #
                # # 提取 crc
                # crc = 1 if rssi == -201 else 0
                if dl_fn == fn and dl_tsn == tsn:
                    ps_info_list.append([burst_type,sub_chan,fn,tsn,beamid,bandid,crc,dl_caridx])
                else:
                    ps_info_list.append([burst_type, sub_chan, fn, tsn, beamid,bandid,crc, -1])

        psfile_result_dict[ps_file_path] = ps_info_list
        linecache.clearcache()

    return psfile_result_dict


if __name__ == '__main__':
    time1 = time.time()
    ps_file_paths = find_ps_files_list(r"C:\Users\Lenovo\Desktop\V1.2联调重要日志\20240518-2-石家庄-组网星地联调-密态-上海台-窄带导航信息增强\20240518_152551383")
    ps_info_result_dict = data_deal(ps_file_paths)
    # .xlsx文件保存，所有文件内容存入一个文件中，每个文件的数据一个sheet工作表
    sava_data_xlsx(ps_info_result_dict)
    time2 = time.time()
    print(time2 - time1)




